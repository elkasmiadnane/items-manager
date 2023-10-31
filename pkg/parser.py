import re
import openpyxl as xl

db = './input/db.xlsx'
wb = xl.load_workbook(db)
ws = wb.active

def parseId(name:str):
    id = re.sub("\.","",
            re.sub("-","",
                re.sub(" ","",
                       re.sub("_","",name))))
    return id
def dbToDict():
    dbObject = {}
    counter = 0
    for index,row in enumerate(ws):
        #print (index)
        name = row[0].value
        id = parseId(str(name))
        exists = row[1].value
        link = row[2].value
        status = row[3].value
        if index > 0:
            if id not in dbObject.keys():
                counter = counter + 1
                dbObject[id] = {"index":int(counter),
                                "id" : id,
                               "name" : name ,
                               "exists" : exists ,
                               "link" : link ,
                               "status" : status}



    return dbObject
class DB():

    def getDB(self):
        database = dbToDict()
        return database

    def getFilterVals(self,arg):
        return {
                'vals'  :("exists" ,"status"),
                'exists':["Yes","No"],
                'status':["ok","None"]
                }

    def insert_row(data:dict):
        rowIndex = int(ws.max_row+1)
        ws.insert_rows(rowIndex)
        for col, value in enumerate(data.values()):
            ws.cell(row=rowIndex, column=col+1, value= '' if value is None else str(value))
            print(str(value))

        wb.save(db)

    def modify_row(idx : int , data: dict):
        for col, value in enumerate(data.values()):
            if col>1:
                cl = ws.cell(row=idx, column=col )
                cl.value = '' if (value is None and str(value).lower() != "none") else str(value)
                print(col +1,str(value) , " ----- ",data.values())

        wb.save(db)

    def delete_row(row:int):
        ws.delete_rows(row,1)